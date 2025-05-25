import os
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import ttk, messagebox, filedialog
import openpyxl

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Config Role Staff")
        self.geometry("700x450")
        self.resizable(True, True)
        
        # Create frame
        self.frame = ttk.Frame(self)
        self.frame.pack(pady=40)

        # Initialize tree as None - will be created when Excel file is loaded
        self.tree = None
        
        # Buttons frame
        self.button_frame = ttk.Frame(self)
        self.button_frame.pack(pady=10)

        # Buttons
        self.refresh_button = ttk.Button(self.button_frame, text='Refresh Data', command=self.button_clicked)
        self.refresh_button.pack(side="left", padx=5)
        
        self.modify_button = ttk.Button(self.button_frame, text='Modify Selected', command=self.modify_row)
        self.modify_button.pack(side="left", padx=5)

        self.open_button = ttk.Button(self.button_frame, text='Open File', command=self.add_excel)
        self.open_button.pack(side="left", padx=5)

        self.selected_item = None

    def create_tree(self, columns):
        # If tree exists, destroy it
        if self.tree is not None:
            self.tree.destroy()
            
        # Create new Treeview with dynamic columns
        self.tree = ttk.Treeview(self.frame, columns=columns, show="headings", height=10)
        
        # Define columns dynamically
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=200)  # Adjust width as needed
        
        # Add scrollbar
        self.scrollbar = ttk.Scrollbar(self.frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.scrollbar.set)
        
        # Pack widgets
        self.tree.pack(side="left")
        self.scrollbar.pack(side="right", fill="y")
        
        # Bind select event
        self.tree.bind('<<TreeviewSelect>>', self.item_selected)

    def item_selected(self, event):
        self.selected_item = self.tree.selection()

    def button_clicked(self):
        if self.tree is None:
            messagebox.showwarning("Warning", "Please load an Excel file first")
            return
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        # Reload data
        self.add_excel()
        messagebox.showinfo(title='Information', message='Data refreshed!')

    def add_excel(self):
        try:
            self.filetype = (("Excel files","*.xlsx"),("All files","*.*"))
            self.filename = filedialog.askopenfilename(title='Open a file', filetypes=self.filetype)
            if not self.filename:  # User cancelled file selection
                return
                
            get_file = os.path.abspath(self.filename)
            wb = openpyxl.load_workbook(get_file)
            sheet = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in sheet[1]]
            
            # Create new tree with these headers
            self.create_tree(headers)
            
            # Add data from remaining rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", "end", values=row)
                
            wb.close()
            
        except Exception as e:
            messagebox.showinfo(title='Error', message=f'Error loading Excel file: {str(e)}')

    def modify_row(self):
        if self.tree is None:
            messagebox.showwarning("Warning", "Please load an Excel file first")
            return
            
        if not self.selected_item:
            messagebox.showwarning("Warning", "Please select a row to modify")
            return
        
        # Get selected item values
        values = self.tree.item(self.selected_item)['values']
        
        # Create modification window
        self.modify_window = tk.Toplevel(self)
        self.modify_window.title("Modify Row")
        self.modify_window.geometry("400x350")
        
        # Entry variables
        self.entries = {}
        self.vars = {}
        
        # Get column headers
        columns = self.tree['columns']
        
        # Create entry fields
        for i, col in enumerate(columns):
            label = ttk.Label(self.modify_window, text=f"{col}:")
            label.grid(row=i, column=0, padx=5, pady=5, sticky="e")
            
            self.vars[col] = tk.StringVar(value=values[i])
            entry = ttk.Entry(self.modify_window, textvariable=self.vars[col])
            entry.grid(row=i, column=1, padx=5, pady=5, sticky="w")
            self.entries[col] = entry

        # Save button
        save_btn = ttk.Button(self.modify_window, text="Save", command=self.save_modifications)
        save_btn.grid(row=len(columns), column=0, columnspan=2, pady=20)

    def save_modifications(self):
        try:
            # Get column headers
            columns = self.tree['columns']
            
            # Get values from entries
            new_values = [self.vars[col].get() for col in columns]
            
            # Update Excel file
            wb = openpyxl.load_workbook(self.filename)
            sheet = wb.active
            
            # Find the row in Excel and update it
            old_values = self.tree.item(self.selected_item)['values']
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if str(row[0].value) == str(old_values[0]):  # Match by first column
                    for i, cell in enumerate(row):
                        cell.value = new_values[i]
                    break
            
            # Save Excel file
            wb.save(self.filename)
            
            # Update Treeview
            self.tree.item(self.selected_item, values=new_values)
            
            # Close modification window
            self.modify_window.destroy()
            messagebox.showinfo("Success", "Row updated successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update row: {str(e)}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
