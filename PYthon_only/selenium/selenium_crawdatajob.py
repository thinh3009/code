import unittest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.webdriver.edge.options import Options
import requests
import time
class PythonOrgSearch():
    def __init__(self):
        options = Options()
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--start-maximized")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)

        self.driver = webdriver.Edge(options=options)
        self.jobs = []

    def test_search(self):
        try:
            self.driver.get("https://www.topcv.vn/viec-lam")
            
        except Exception as e:
            print("Error accessing page:", e)
            return  # Dừng nếu không vào được

        try:
            keyword_input = self.driver.find_element(By.ID, "keyword")
            keyword_input.send_keys("software engineer")
            
            search_button = self.driver.find_element(By.XPATH, "//*[@id='frm-search-job']/div/div[4]/button")
            search_button.click()

            time.sleep(30)
        except Exception as e:
            print("Error interacting with search page:", e)

    def scroll_page(self, scroll_value):
        self.driver.execute_script(f"window.scrollBy(0, {scroll_value})")
        time.sleep(10)
    def find(self):
        self.htmltxt = requests.get('https://www.topcv.vn/tim-viec-lam-software-engineer?sba=1').text
        self.soup = BeautifulSoup(self.htmltxt, 'lxml')
        self.detail = self.soup.find_all('div', class_='job-item-search-result')
        
        # Extract all job information
        for job in self.detail:
            title = job.find('h3').text.strip()
            company = job.find('a', class_='company').span.text.strip() if job.find('a', class_='company') else 'N/A'
            location = job.find('div', class_='label-content').span.text.strip() if job.find('div', class_='label-content') else 'N/A'
            salary = job.find('div', class_='box-right').label.text.strip() if job.find('div', class_='box-right') else 'N/A'
            url = job.find('h3', class_='title').a.get('href') if job.find('h3', class_='title') else 'N/A'
            companyInfor=job.find('a', class_='company').get('href') if job.find('a', class_='company') else 'N/A'
            
            self.jobs.append({
                'Title': title,
                'Company': company,
                'Location': location,
                'Salary': salary,
                'URL': url,
                'Company information': companyInfor
            })

    def save_to_excel(self, filename='jobs.xlsx'):
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # Write headers (first row)
        headers = ['Title', 'Company', 'Location', 'Salary']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Write job data
        for row, job in enumerate(self.jobs, 2):
            sheet.cell(row=row, column=1, value=job['Title'])
            sheet.cell(row=row, column=2, value=job['Company'])
            sheet.cell(row=row, column=3, value=job['Location'])
            sheet.cell(row=row, column=4, value=job['Salary'])
        
        # Save the workbook
        wb.save(filename)

    def find_title(self):
        if not self.jobs:  # If jobs haven't been scraped yet
            self.find()
        
        for job in self.jobs:
            print(f"Title: {job['Title']}")
            print(f"Company: {job['Company']}")
            print(f"Location: {job['Location']}")
            print(f"Salary: {job['Salary']}")
            print(f"URL: {job['URL']}")
            print(f"Company information: {job['Company information']}")
            print("-" * 50)

    def tearDown(self):
        self.driver.close()
    

if __name__ == "__main__":
    scroll=9500
    test_find=PythonOrgSearch()
    test_find.test_search()
    test_find.find_title()
    test_find.scroll_page(scroll)
    
